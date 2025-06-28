# excel_writer.py

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from config import EXCEL_REPORT_DIR

# Карта действий → колонка
ACTION_COLUMN_MAP = {
    "Пришел на работу": 2,
    "Вышел на обед": 3,
    "Вернулся с обеда": 4,
    "Ушел с работы": 5
}

# Основная функция записи события в Excel
def write_event_to_excel(user_id, username, action, timestamp_str):
    timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
    month_name = timestamp.strftime('%B')  # Название месяца (напр. 'June')
    date_str = timestamp.strftime('%Y-%m-%d')
    time_str = timestamp.strftime('%H:%M:%S')

    # Создание директории, если не существует
    os.makedirs(EXCEL_REPORT_DIR, exist_ok=True)
    file_path = os.path.join(EXCEL_REPORT_DIR, f"report_{user_id}_2025_by_months.xlsx")


    wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
    if not os.path.exists(file_path):
        wb.remove(wb.active)  # удаляем дефолтный лист

    if month_name not in wb.sheetnames:
        ws = wb.create_sheet(title=month_name)
        ws.append(["Дата", "Начало", "Обед‑выход", "Обед‑возврат", "Конец"])
    else:
        ws = wb[month_name]

    # Поиск строки по дате
    row_found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == date_str:
            col = ACTION_COLUMN_MAP.get(action)
            if col:
                ws.cell(row=row, column=col).value = time_str
            row_found = True
            break

    # Если дата не найдена — создаём новую строку
    if not row_found:
        new_row = [date_str, "", "", "", ""]
        col_index = ACTION_COLUMN_MAP.get(action)
        if col_index:
            new_row[col_index - 1] = time_str
        ws.append(new_row)

    # Применяем стили к таблице
    for col in ws.columns:
        for cell in col:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center')
            if cell.row == 1:
                cell.font = Font(bold=True)

    # Сохраняем файл
    wb.save(file_path)
    print(f"[DEBUG] Записан {action} {date_str} {time_str} в файл {file_path}")
