import os
import csv
from datetime import datetime, time
from zoneinfo import ZoneInfo
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from storage import WORKTIME_FILE  


TS_ZONE = ZoneInfo("Asia/Bishkek")

START_TIME = time(8, 30)


def generate_excel_report_by_months(user_id, username, today_only=False):

    if not os.path.exists(WORKTIME_FILE):
        print(f"[ERROR] Файл {WORKTIME_FILE} не найден.")
        return None


    data_by_month = {m: {} for m in range(1, 13)}
    with open(WORKTIME_FILE, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 3:
                continue
            uid, action, ts = row[0], row[1], row[-1]
            if uid != str(user_id):
                continue
            try:
                dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
            d = dt.date()
            m = d.month
            rec = data_by_month[m].setdefault(
                d, {"start": None, "lunch_out": None, "lunch_in": None, "end": None}
            )
            if action == "Пришел на работу":
                rec["start"] = dt
            elif action == "Вышел на обед":
                rec["lunch_out"] = dt
            elif action == "Вернулся с обеда":
                rec["lunch_in"] = dt
            elif action == "Ушел с работы":
                rec["end"] = dt


    if today_only:
        today = datetime.now(TS_ZONE).date()
        m = today.month
        rec = data_by_month.get(m, {}).get(
            today, {"start": None, "lunch_out": None, "lunch_in": None, "end": None}
        )
        data_by_month = {m: {today: rec}}


    wb = Workbook()
    wb.remove(wb.active)

    total_minutes = total_late = total_days = total_absent = 0
    now = datetime.now(TS_ZONE)

    for month in range(1, 13):
        records = data_by_month.get(month, {})

        ws = wb.create_sheet(title=datetime(2025, month, 1).strftime('%B'))


        ws.append(["Дата", "Начало", "Обед-выход", "Обед-возврат", "Конец", "Часы", "Опоздание"])


        month_minutes = 0

        for date_key, times in sorted(records.items()):
            st = times["start"]
            lo = times["lunch_out"]
            li = times["lunch_in"]
            en = times["end"]

            minutes = 0
            late_flag = ""

            if st:

                lo_eff = lo or now
                morning = max(int((lo_eff - st).total_seconds() // 60), 0)


                if li:
                    en_eff = en or now
                    afternoon = max(int((en_eff - li).total_seconds() // 60), 0)
                else:
                    afternoon = 0

                minutes = morning + afternoon
                total_days += 1


                if st.time() > START_TIME:
                    late_flag = "✅"
                    total_late += 1
            else:
                total_absent += 1

            total_minutes += minutes
            month_minutes += minutes

            hours = round(minutes / 60, 2) if minutes else ''
            ws.append([
                date_key.strftime('%Y-%m-%d'),
                st.strftime('%H:%M:%S') if st else '',
                lo.strftime('%H:%M:%S') if lo else '',
                li.strftime('%H:%M:%S') if li else '',
                en.strftime('%H:%M:%S') if en else '',
                hours,
                late_flag
            ])


        ws.append([])

        month_hours = round(month_minutes / 60, 2)
        ws.append(['Итого часов за месяц:', month_hours])


        for col in ws.columns:
            for cell in col:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
                if cell.row == 1:
                    cell.font = Font(bold=True)


    total_hours = round(total_minutes / 60, 2)
    summary = wb.create_sheet(title="Итоги")
    summary.append(["Пользователь", "ID", "Всего часов", "Рабочих дней", "Опозданий", "Пропусков"])
    summary.append([username, user_id, total_hours, total_days, total_late, total_absent])


    norm = wb.create_sheet(title="Норма 2025")
    norm.append(["Месяц", "Норма часов"])
    for m in range(1, 13):
        norm.append([datetime(2025, m, 1).strftime('%B'), ""])


    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    print(f"[DEBUG] Отчёт сформирован: {username} ({user_id})")
    return buf
