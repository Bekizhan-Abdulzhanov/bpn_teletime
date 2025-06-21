import telebot
from datetime import datetime
import csv
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton,ReplyKeyboardMarkup, KeyboardButton
from apscheduler.schedulers.background import BackgroundScheduler
import time
import pytz
import qrcode
from waitress import serve
from flask import Flask
from dotenv import load_dotenv
load_dotenv()
import calendar



import os

os.environ['TZ'] = 'Asia/Bishkek'

kyrgyzstan_tz = pytz.timezone('Asia/Bishkek')
now = datetime.now(kyrgyzstan_tz)

TOKEN = os.environ.get("TOKEN")
if not TOKEN:
    raise Exception('Переменная окружения TOKEN не установлена')

bot = telebot.TeleBot(TOKEN)

app = Flask(__name__)

@app.route('/')
def index():
    return "Bot is running."


def run_flask():
    serve(app, host='0.0.0.0', port=5000)


common_qr_url = "https://t.me/BPN_KG_managetime_bot?start=checkin"
qr = qrcode.make(common_qr_url)
qr.save("common_qr_code.png")


WORKTIME_FILE = 'work_time.csv'
EXCEL_REPORT_DIR = 'work_reports'
USERS_FILE = 'users.csv'
ADMIN_ID = 557174721

  



if not os.path.exists(EXCEL_REPORT_DIR):
    os.makedirs(EXCEL_REPORT_DIR)



def save_work_time(user_id, user_name, action): 
    now = datetime.now(kyrgyzstan_tz).strftime("%Y-%m-%d %H:%M:%S")  
    file_exists = os.path.isfile(WORKTIME_FILE) 

    try: 
        with open(WORKTIME_FILE, mode='a', newline='', encoding='utf-8') as file: 
            writer = csv.writer(file) 
            if not file_exists: 
                writer.writerow(["User ID", "User Name", "Action", "Timestamp"]) 
            writer.writerow([user_id, user_name, action, now]) 
        print(f"[{now}] Записано: {user_name} ({user_id}) - {action}") 
    except Exception as e: 
        print(f"Ошибка при записи времени: {e}") 


now = datetime.now(kyrgyzstan_tz).strftime("%Y-%m-%d %H:%M:%S")

scheduler = BackgroundScheduler()


@bot.message_handler(commands=['register'])
def register(message):
    user = message.from_user
    with open(USERS_FILE, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow([user.id, user.username, 'pending'])
    bot.send_message(ADMIN_ID, f"Новая заявка на регистрацию от {user.username} ({user.id})",
                     reply_markup=approve_keyboard(user.id))
    bot.reply_to(message, "Ваша заявка отправлена администратору на рассмотрение.")


def approve_keyboard(user_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("Принять", callback_data=f"approve_{user_id}"))
    markup.add(InlineKeyboardButton("Отклонить", callback_data=f"reject_{user_id}"))
    return markup

@bot.callback_query_handler(func=lambda call: call.data.startswith("approve_"))
def approve_user(call):
    user_id = call.data.split("_")[1]
    update_user_status(user_id, "approved")
    bot.send_message(user_id, "Ваша регистрация одобрена!")
    bot.send_message(ADMIN_ID, f"Пользователь {user_id} одобрен.")

@bot.callback_query_handler(func=lambda call: call.data.startswith("reject_"))
def reject_user(call):
    user_id = call.data.split("_")[1]
    update_user_status(user_id, "rejected")
    bot.send_message(user_id, "Ваша заявка отклонена.")
    bot.send_message(ADMIN_ID, f"Пользователь {user_id} отклонен.")


def update_user_status(user_id, status):
    users = []
    with open(USERS_FILE, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if row[0] == user_id:
                row[2] = status
            users.append(row)
    with open(USERS_FILE, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerows(users)


def is_user_approved(user_id):
    with open(USERS_FILE, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if row[0] == str(user_id) and row[2] == 'approved':
                return True
    return False



@bot.message_handler(commands=['1'])
def show_menu(message):
    if not is_user_approved(message.from_user.id):
        bot.reply_to(message, "Вы не зарегистрированы или ваша заявка не одобрена.")
        return

    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    markup.add(
        KeyboardButton("🍽 Вышел на обед")
    )
    markup.add(
        KeyboardButton("🍽 Вернулся с обеда"),
        KeyboardButton("🏁 Ушел с работы")
    )
    bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)


@bot.message_handler(func=lambda message: message.text in [ "🍽 Вышел на обед", "🍽 Вернулся с обеда", "🏁 Ушел с работы"])
def handle_work_time(message):
    if not is_user_approved(message.from_user.id):
        bot.reply_to(message, "Вы не зарегистрированы или ваша заявка не одобрена.")
        return

    user = message.from_user
    actions = {
        "🍽 Вышел на обед": "Вышел на обед",
        "🍽 Вернулся с обеда": "Вернулся с обеда",
        "🏁 Ушел с работы": "Ушел с работы"
    }
    
    action = actions[message.text]
    save_work_time(user.id, user.username, action)
    bot.reply_to(message, f"Вы отметили: {message.text}")



@bot.message_handler(commands=['start'])
def start_work(message):
    if not is_user_approved(message.from_user.id):
        bot.reply_to(message, "Вы не зарегистрированы или ваша заявка не одобрена.")
        return
    user = message.from_user
    save_work_time(user.id, user.username, "Пришел на работу")
    bot.reply_to(message, f"Доброе утро, {user.first_name}! Вы отметили приход на работу.")

@bot.message_handler(commands=['lunch_out'])
def lunch_out(message):
    if not is_user_approved(message.from_user.id):
        bot.reply_to(message, "Вы не зарегистрированы или ваша заявка не одобрена.")
        return    
    user = message.from_user
    save_work_time(user.id, user.username, "Вышел на обед")
    bot.reply_to(message, "Вы отметили выход на обед.")

@bot.message_handler(commands=['lunch_in'])
def lunch_in(message):
    if not is_user_approved(message.from_user.id):
        bot.reply_to(message, "Вы не зарегистрированы или ваша заявка не одобрена.")
        return    
    user = message.from_user
    save_work_time(user.id, user.username, "Вернулся с обеда")
    bot.reply_to(message, "Вы отметили возвращение с обеда.")

@bot.message_handler(commands=['end'])
def end_work(message):
    if not is_user_approved(message.from_user.id):
        bot.reply_to(message, "Вы не зарегистрированы или ваша заявка не одобрена.")
        return     
    user = message.from_user
    save_work_time(user.id, user.username, "Ушел с работы")
    bot.reply_to(message, f"До свидания, {user.first_name}! Вы отметили уход с работы.")

@bot.message_handler(commands=['ping'])
def ping(message):
    bot.reply_to(message, "pong")

@bot.message_handler(commands=['all_reports'])
def all_reports(message):
    if message.from_user.id != ADMIN_ID:
        bot.reply_to(message, "У вас нет доступа к этому разделу.")
        return
    for user_id in get_all_users():
        report_file = generate_excel_report(user_id)
        if report_file:
            with open(report_file, 'rb') as file:
                bot.send_document(ADMIN_ID, file, caption=f"Отчет сотрудника {user_id}")

@bot.message_handler(commands=['edit_time'])
def edit_time(message):
    if message.from_user.id != ADMIN_ID:
        bot.reply_to(message, "У вас нет прав для редактирования данных.")
        return
    
    markup = InlineKeyboardMarkup()
    
    if not os.path.exists(WORKTIME_FILE):
        bot.reply_to(message, "Файл с рабочим временем пуст или не существует.")
        return
    
    users = {}
    with open(WORKTIME_FILE, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader, None)  
        for row in reader:
            if len(row) < 2:
                continue
            user_id, user_name = row[0], row[1]
            users[user_id] = user_name

    for user_id, user_name in users.items():
        markup.add(InlineKeyboardButton(text=user_name, callback_data=f"edit_{user_id}"))

    bot.send_message(message.chat.id, "Выберите сотрудника для редактирования:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("edit_"))
def select_employee(call):
    user_id = call.data.split("_")[1]
    
    markup = InlineKeyboardMarkup()
    
    dates = set()
    with open(WORKTIME_FILE, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader, None)
        for row in reader:
            if len(row) < 4:
                continue
            if row[0] == user_id:
                date = row[3].split()[0]
                dates.add(date)

    for date in sorted(dates):
        markup.add(InlineKeyboardButton(text=date, callback_data=f"date_{user_id}_{date}"))

    bot.send_message(call.message.chat.id, "Выберите дату для редактирования:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("date_"))
def select_date(call):
    _, user_id, date = call.data.split("_")

    markup = InlineKeyboardMarkup()
    actions = ["Пришел на работу", "Вышел на обед", "Вернулся с обеда", "Ушел с работы"]

    for action in actions:
        markup.add(InlineKeyboardButton(text=f"Изменить {action}", callback_data=f"change_{user_id}_{date}_{action}"))

    bot.send_message(call.message.chat.id, "Выберите действие для редактирования:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("change_"))
def change_time(call):
    _, user_id, date, action = call.data.split("_")

    msg = bot.send_message(call.message.chat.id, f"Введите новое время для '{action}' в формате ЧЧ:ММ:")
    bot.register_next_step_handler(msg, lambda message: update_time(message, user_id, date, action))


def update_time(message, user_id, date, action): 
    new_time = message.text.strip()
    try: 
        datetime.strptime(new_time, "%H:%M") 
    except ValueError: 
        bot.send_message(message.chat.id, "Некорректный формат. Введите время в формате ЧЧ:ММ.") 
        return 
    
    updated_rows = [] 
    updated = False 
    
    with open(WORKTIME_FILE, mode='r', encoding='utf-8') as file: 
        reader = csv.reader(file) 
        header = next(reader) 
        for row in reader: 
            if len(row) < 4: 
                continue 
            
            row_date = row[3].split()[0] 
            if row[0] == user_id and row_date == date and row[2] == action: 
                row[3] = f"{date} {new_time}:00" 
                updated = True 
            updated_rows.append(row) 
    
    if updated: 
        with open(WORKTIME_FILE, mode='w', newline='', encoding='utf-8') as file: 
            writer = csv.writer(file) 
            writer.writerow(["User ID", "Имя пользователя", "Действие", "Время"]) 
            writer.writerows(updated_rows) 

        bot.send_message(message.chat.id, f"Время для '{action}' сотрудника {user_id} на {date} обновлено на {new_time}.") 
        generate_excel_report(user_id)  # Добавлено обновление Excel 
    else: 
        bot.send_message(message.chat.id, "Запись не найдена или уже обновлена.")
    
    # После обновления CSV сразу генерируем новый отчет
    generate_excel_report(user_id)


def get_all_users():
    users = {}
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                users[row[0]] = row[1]  
    return users




def generate_final_monthly_report(user_id, user_name, user_log, output_dir="work_reports"):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb = Workbook()

    for month_idx in range(1, 13):
        month_data = {
            dt: user_log[dt] for dt in user_log
            if dt.month == month_idx
        }
        if not month_data:
            continue

        sheet_name = calendar.month_name[month_idx]
        ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells('A1:I1')
        ws['A1'] = "Табель / Arbeitsblatt"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws['A2'] = "Компания / Firma:"
        ws['B2'] = "ОсОО «Business Academy BPN»"
        ws['A3'] = "Имя сотрудника / Name des Mitarbeiters:"
        ws['B3'] = user_name
        ws['A4'] = "Должность / Funktion:"
        ws['B4'] = "Сотрудник"

        headers = ["День", "Дата", "Часы работы", "Отсутствие", "Переработка"]
        ws.append([])
        ws.append(headers)

        total_hours = 0
        work_norm = 8

        for day in range(1, 32):
            try:
                date = datetime(datetime.now().year, month_idx, day).date()
            except ValueError:
                continue

            record = user_log.get(date, {})
            worked_hours = 0
            if "start" in record and "end" in record:
                worked_hours = (record["end"] - record["start"]).seconds / 3600
                if "lunch_out" in record and "lunch_in" in record:
                    worked_hours -= (record["lunch_in"] - record["lunch_out"]).seconds / 3600
                worked_hours = round(worked_hours, 2)

            total_hours += worked_hours

            ws.append([
                day,
                date.strftime("%d.%m.%Y"),
                worked_hours if worked_hours else "",
                "" if worked_hours else "1",
                ""
            ])

        ws.append([])
        ws.append(["", "", "Итого отработано / Total geleistete Arbeitszeit:", round(total_hours, 2)])
        ws.append(["", "", "Итого отсутствие / Total Absenz:", "0"])
        ws.append(["", "", "Итого переработки / Total Überstunden:", "0"])

        ws.append([])
        ws.append(["Поле для пояснений сотрудника / Notizen und Anmerkungen Mitarbeiter:"])
        ws.append([])
        ws.append(["Поле для пояснений бухгалтера / Notizen und Anmerkungen Buchhaltung:"])
        ws.append([])
        ws.append(["", "", "Переработки на начало / Überstunden aus Vorperiode:", ""])
        ws.append(["", "", "Отработано за месяц / Total geleistete Arbeitszeit IST:", round(total_hours, 2)])
        ws.append(["", "", "Переработки за месяц / Total Überstunden in diesem Monat:", ""])
        ws.append(["", "", "Месячный Норматив / Total Arbeitszeit SOLL:", "152"])
        ws.append(["", "", "Переработка на конец / Total Überstunden Aktuell:", ""])
        ws.append(["", "", "Оплачиваемых часов / Bezahlte Arbeitszeit:", ""])

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=6, max_col=5, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    file_name = f"{output_dir}/BPN_Report_{user_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(file_name)
    return file_name


# --------------------- ДОБАВЛЕННАЯ КОМАНДА ДЛЯ БОТА ---------------------
@bot.message_handler(commands=['report_sheet'])
def send_monthly_sheet(message):
    user_id = message.from_user.id
    user_name = message.from_user.username

    log_data = {}
    if os.path.exists(WORKTIME_FILE):
        with open(WORKTIME_FILE, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader, None)
            for row in reader:
                if len(row) < 4:
                    continue
                uid, uname, action, timestamp = row
                if str(user_id) != uid:
                    continue
                dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                if dt.date() not in log_data:
                    log_data[dt.date()] = {}
                if action == "Пришел на работу":
                    log_data[dt.date()]["start"] = dt
                elif action == "Вышел на обед":
                    log_data[dt.date()]["lunch_out"] = dt
                elif action == "Вернулся с обеда":
                    log_data[dt.date()]["lunch_in"] = dt
                elif action == "Ушел с работы":
                    log_data[dt.date()]["end"] = dt

    if not log_data:
        bot.reply_to(message, "Нет данных для отчета.")
        return

    file_path = generate_final_monthly_report(user_id, user_name, log_data)

    if file_path and os.path.exists(file_path):
        with open(file_path, "rb") as file:
            bot.send_document(message.chat.id, file, caption="Ваш табель по месяцам (Excel).")
    else:
        bot.reply_to(message, "Не удалось сгенерировать отчет.")



def run_bot():
    bot.infinity_polling()


flask_thread = threading.Thread(target=run_flask)
flask_thread.start()


def run_bot():
    bot.infinity_polling()

if __name__ == '__main__':
    
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()

    bot.remove_webhook()

    bot.infinity_polling()

PORT =int(os.environ.get("PORT",5000))
def run_flask():
    serve(app,host='0.0.0.0',port=PORT)

import datetime
import time
from apscheduler.schedulers.background import BackgroundScheduler

scheduler = BackgroundScheduler()

AUTO_USERS = {
    378268765: "ErlanNasiev",
    557174721: "BekizhanAbdulzhanov",
}

# Настроим расписание работы
def schedule_auto_records():
    for user_id, username in AUTO_USERS.items():
        # ПН, СР, ПТ - 08:29, 12:00, 13:00, 17:30
        scheduler.add_job(save_work_time, "cron", day_of_week="mon,wed,fri", hour=8, minute=29,
                          args=[user_id, username, "Пришел на работу"])
        scheduler.add_job(save_work_time, "cron", day_of_week="mon,wed,fri", hour=12, minute=0,
                          args=[user_id, username, "Вышел на обед"])
        scheduler.add_job(save_work_time, "cron", day_of_week="mon,wed,fri", hour=13, minute=0,
                          args=[user_id, username, "Вернулся с обеда"])
        scheduler.add_job(save_work_time, "cron", day_of_week="mon,wed,fri", hour=17, minute=30,
                          args=[user_id, username, "Ушел с работы"])

        # ВТ, ЧТ - 08:28, 12:00, 13:00, 17:30
        scheduler.add_job(save_work_time, "cron", day_of_week="tue,thu", hour=8, minute=28,
                          args=[user_id, username, "Пришел на работу"])
        scheduler.add_job(save_work_time, "cron", day_of_week="tue,thu", hour=12, minute=0,
                          args=[user_id, username, "Вышел на обед"])
        scheduler.add_job(save_work_time, "cron", day_of_week="tue,thu", hour=13, minute=0,
                          args=[user_id, username, "Вернулся с обеда"])
        scheduler.add_job(save_work_time, "cron", day_of_week="tue,thu", hour=17, minute=30,
                          args=[user_id, username, "Ушел с работы"])


schedule_auto_records()

time.sleep(2) 
if not scheduler.running:
    scheduler.start()

print("Автоматическая запись времени включена.")
print("Текущее время:", datetime.datetime.now())

